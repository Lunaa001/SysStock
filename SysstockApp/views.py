# =========================
# IMPORTS
# =========================
from rest_framework import viewsets, permissions, filters, status
from rest_framework.exceptions import PermissionDenied
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.response import Response

from django.db import transaction
from django.utils import timezone
from django.utils.timezone import localdate
from datetime import datetime, time

from openpyxl import Workbook
from django.http import HttpResponse

from django.db.models import Sum

from .models import Category, Branch, Product, StockMovement, Sale
from .serializers import (
    CategorySerializer,
    BranchSerializer,
    ProductSerializer,
    StockMovementSerializer,
    SaleSerializer,
)
from AccountAdmin.permissions import IsAdmin  # alias válido a IsAdminRole


# =========================
# HELPERS DE TENANT / SCOPE
# =========================
def _scope_branches(qs, user):
    """
    - superuser: ve todo
    - admin: ve solo sucursales donde es owner
    - limMerchant: solo su sucursal asignada
    """
    if getattr(user, "is_superuser", False):
        return qs
    if getattr(user, "rol", None) == "admin":
        return qs.filter(owner=user)
    # limMerchant
    return qs.filter(pk=user.sucursal_id)


def _scope_by_branch_on_model(qs, user, branch_field="sucursal"):
    """
    Para modelos con FK a sucursal (p.ej., Product.sucursal, Sale.sucursal):
    - superuser: todo
    - admin: {branch_field}__owner = user
    - limMerchant: {branch_field} = user.sucursal
    """
    if getattr(user, "is_superuser", False):
        return qs
    if getattr(user, "rol", None) == "admin":
        return qs.filter(**{f"{branch_field}__owner": user})
    return qs.filter(**{f"{branch_field}__id": user.sucursal_id})


# =========================
# HELPER DE STOCK (por movimientos)
# =========================
def _stock_actual_producto(producto: Product) -> int:
    """
    Calcula stock actual por producto como:
    stock = ENTRADAS ('IN') - SALIDAS ('OUT')
    """
    entradas = StockMovement.objects.filter(producto=producto, tipo="IN").aggregate(total=Sum("cantidad"))["total"] or 0
    salidas  = StockMovement.objects.filter(producto=producto, tipo="OUT").aggregate(total=Sum("cantidad"))["total"] or 0
    return int(entradas - salidas)


# =========================
# SUCURSALES
# =========================
class BranchViewSet(viewsets.ModelViewSet):
    serializer_class = BranchSerializer
    permission_classes = [permissions.IsAuthenticated]
    queryset = Branch.objects.all().order_by("id")

    def get_queryset(self):
        qs = Branch.objects.all().order_by("id")
        return _scope_branches(qs, self.request.user)

    def perform_create(self, serializer):
        """
        Guarda owner=admin que crea la sucursal (si no, luego no aparece en /api/sucursales/ del admin).
        """
        user = self.request.user
        if getattr(user, "rol", None) != "admin" and not getattr(user, "is_superuser", False):
            raise PermissionDenied("Solo un admin puede crear sucursales.")
        serializer.save(owner=user)

    # ✅ EDITAR SUCURSAL (PUT/PATCH): nombre, dirección y teléfono
    def update(self, request, *args, **kwargs):
        branch = self.get_object()
        user = request.user
        if branch.owner != user and not getattr(user, "is_superuser", False):
            return Response({"detail": "No puedes editar una sucursal de otro admin."}, status=403)
        return super().update(request, *args, **kwargs)

    def partial_update(self, request, *args, **kwargs):
        branch = self.get_object()
        user = request.user
        if branch.owner != user and not getattr(user, "is_superuser", False):
            return Response({"detail": "No puedes editar una sucursal de otro admin."}, status=403)
        return super().partial_update(request, *args, **kwargs)

    # -------------------------
    # /api/sucursales/<id>/ventas_rango/?desde=YYYY-MM-DD&hasta=YYYY-MM-DD
    # -------------------------
    @action(detail=True, methods=["get"], url_path="ventas_rango")
    def ventas_rango(self, request, pk=None):
        branch = self.get_object()
        user = request.user

        # Permisos
        if getattr(user, "rol", None) == "limMerchant" and user.sucursal_id != branch.id:
            return Response({"detail": "No tienes permiso para ver esta sucursal."}, status=403)
        if getattr(user, "rol", None) == "admin" and branch.owner_id != user.id:
            return Response({"detail": "Esta sucursal no pertenece a tu cuenta."}, status=403)

        # Parámetros requeridos
        desde_str = request.query_params.get("desde")
        hasta_str = request.query_params.get("hasta")
        if not desde_str or not hasta_str:
            return Response(
                {"detail": "Parámetros requeridos: desde=YYYY-MM-DD & hasta=YYYY-MM-DD"},
                status=400
            )

        # Parseo correcto de fechas
        try:
            desde_date = datetime.strptime(desde_str, "%Y-%m-%d").date()
            hasta_date = datetime.strptime(hasta_str, "%Y-%m-%d").date()
        except ValueError:
            return Response({"detail": "Formato inválido. Usa YYYY-MM-DD."}, status=400)

        if hasta_date < desde_date:
            return Response({"detail": "`hasta` no puede ser anterior a `desde`."}, status=400)

        # Rango completo por día [00:00:00, 23:59:59]
        desde_dt = datetime.combine(desde_date, time.min)
        hasta_dt = datetime.combine(hasta_date, time.max)

        # Ventas del rango
        ventas = (
            Sale.objects.filter(sucursal=branch, creado_en__range=(desde_dt, hasta_dt))
            .prefetch_related("items__producto")
            .select_related("sucursal")
            .order_by("creado_en")
        )

        def total_venta(venta):
            return sum(it.cantidad * it.precio_unit for it in venta.items.all())

        monto_total = 0.0
        data = []

        for v in ventas:
            tv = float(total_venta(v))
            monto_total += tv
            data.append({
                "venta_id": v.id,
                "fecha_hora": v.creado_en,
                "total_venta": tv,
                "items": [
                    {
                        "producto": it.producto.nombre if it.producto else None,
                        "cantidad": it.cantidad,
                        "precio_unit": float(it.precio_unit),
                        "total_item": float(it.cantidad * it.precio_unit)
                    } for it in v.items.all()
                ]
            })

        return Response({
            "sucursal": branch.name,
            "desde": desde_str,
            "hasta": hasta_str,
            "monto_total": float(monto_total),
            "ventas": data
        })

    # -------------------------
    # /api/sucursales/<id>/ventas_por_producto/?desde=&hasta=
    # -------------------------
    @action(detail=True, methods=["get"], url_path="ventas_por_producto")
    def ventas_por_producto(self, request, pk=None):
        branch = self.get_object()
        u = request.user
        if getattr(u, "rol", None) == "limMerchant" and u.sucursal_id != branch.id:
            return Response({"detail": "No tienes permiso para esta sucursal."}, status=403)
        if getattr(u, "rol", None) == "admin" and branch.owner_id != u.id:
            return Response({"detail": "Esta sucursal no te pertenece."}, status=403)

        desde = request.query_params.get("desde")
        hasta = request.query_params.get("hasta")
        qs = Sale.objects.filter(sucursal=branch).prefetch_related("items__producto")
        if desde and hasta:
            qs = qs.filter(creado_en__date__range=(desde, hasta))

        agg = {}
        for v in qs:
            for it in v.items.all():
                pid = getattr(it.producto, "id", None)
                pname = getattr(it.producto, "nombre", None)
                if pid is None:
                    continue
                if pid not in agg:
                    agg[pid] = {"producto_id": pid, "producto": pname, "cantidad": 0}
                agg[pid]["cantidad"] += it.cantidad

        result = sorted(agg.values(), key=lambda x: x["cantidad"], reverse=True)
        return Response({
            "sucursal": branch.name,
            "desde": desde,
            "hasta": hasta,
            "resumen": result,
        })

    # -------------------------
    # /api/sucursales/<id>/ventas_por_dia/?desde=&hasta=
    # -------------------------
    @action(detail=True, methods=["get"], url_path="ventas_por_dia")
    def ventas_por_dia(self, request, pk=None):
        branch = self.get_object()
        u = request.user
        if getattr(u, "rol", None) == "limMerchant" and u.sucursal_id != branch.id:
            return Response({"detail": "No tienes permiso para esta sucursal."}, status=403)
        if getattr(u, "rol", None) == "admin" and branch.owner_id != u.id:
            return Response({"detail": "Esta sucursal no te pertenece."}, status=403)

        desde = request.query_params.get("desde")
        hasta = request.query_params.get("hasta")
        qs = Sale.objects.filter(sucursal=branch)
        if desde and hasta:
            qs = qs.filter(creado_en__date__range=(desde, hasta))

        # sumar total por día
        series = {}
        for v in qs.prefetch_related("items"):
            d = v.creado_en.date()
            tot = sum(it.cantidad * it.precio_unit for it in v.items.all())
            series[d] = series.get(d, 0.0) + float(tot)

        rows = [{"fecha": str(k), "monto": v} for k, v in sorted(series.items())]
        total = sum(x["monto"] for x in rows)

        return Response({
            "sucursal": branch.name,
            "desde": desde,
            "hasta": hasta,
            "monto_total": total,
            "serie": rows,
        })

    # -------------------------
    # /api/sucursales/<id>/ventas_export/xlsx/?desde=&hasta=
    # -------------------------
    @action(detail=True, methods=["get"], url_path="ventas_export/xlsx")
    def ventas_export_xlsx(self, request, pk=None):
        branch = self.get_object()
        u = request.user

        # Permisos
        if getattr(u, "rol", None) == "limMerchant" and u.sucursal_id != branch.id:
            return Response({"detail": "No tienes permiso para esta sucursal."}, status=403)
        if getattr(u, "rol", None) == "admin" and branch.owner_id != u.id:
            return Response({"detail": "Esta sucursal no te pertenece."}, status=403)

        # Filtros opcionales de fecha
        desde = request.query_params.get("desde")
        hasta = request.query_params.get("hasta")

        qs = (
            Sale.objects
            .filter(sucursal=branch)
            .prefetch_related("items__producto")
            .select_related("sucursal")
            .order_by("creado_en")
        )
        if desde and hasta:
            qs = qs.filter(creado_en__date__range=(desde, hasta))

        # Armar Excel
        wb = Workbook()
        ws = wb.active
        ws.title = f"Ventas {branch.name}"
        ws.append(["ID Venta", "Fecha/Hora", "Sucursal", "Producto", "Cantidad", "Precio Unit.", "Total Item"])

        for sale in qs:
            for it in sale.items.all():
                ws.append([
                    sale.id,
                    timezone.localtime(sale.creado_en).strftime("%Y-%m-%d %H:%M"),
                    branch.name,
                    getattr(it.producto, "nombre", None),
                    int(it.cantidad),
                    float(it.precio_unit),
                    float(it.cantidad * it.precio_unit),
                ])

        resp = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        filename = f'ventas_{branch.id}.xlsx'
        resp['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(resp)
        return resp

    # -------------------------
    # /api/sucursales/<id>/resumen/?threshold=5&limit=50
    # -------------------------
    @action(detail=True, methods=["get"], url_path="resumen")
    def resumen(self, request, pk=None):
        """
        Resumen reducido (usa stock REAL por movimientos, no Product.cantidad).
        - ventas_hoy.monto (con filtro por rango horario local -> evita problemas de timezone)
        - productos_bajo_stock (stock real calculado; usa stock_min si existe)
        """
        branch = self.get_object()
        u = request.user

        # Permisos
        if getattr(u, "rol", None) == "limMerchant" and u.sucursal_id != branch.id:
            return Response({"detail": "No tienes permiso para esta sucursal."}, status=403)
        if getattr(u, "rol", None) == "admin" and branch.owner_id != u.id:
            return Response({"detail": "Esta sucursal no te pertenece."}, status=403)

        # Hoy local -> rango horario aware [00:00:00, 23:59:59]
        hoy = localdate()
        inicio = timezone.make_aware(datetime.combine(hoy, time.min))
        fin    = timezone.make_aware(datetime.combine(hoy, time.max))

        # Ventas HOY (monto)
        ventas_hoy_qs = (
            Sale.objects
            .filter(sucursal=branch, creado_en__range=(inicio, fin))
            .prefetch_related("items")
        )
        ventas_hoy_monto = float(sum(
            sum(float(it.cantidad) * float(it.precio_unit) for it in v.items.all())
            for v in ventas_hoy_qs
        ))

        # Threshold de bajo stock
        try:
            threshold = int(request.query_params.get("threshold", 5))
        except ValueError:
            threshold = 5

        # Stock bajo (por movimientos)
        productos_bajo_stock = []
        for p in Product.objects.filter(sucursal=branch).select_related("categoria"):
            # stock actual por movimientos
            entradas = StockMovement.objects.filter(producto=p, tipo="IN").aggregate(s=Sum("cantidad"))["s"] or 0
            salidas  = StockMovement.objects.filter(producto=p, tipo="OUT").aggregate(s=Sum("cantidad"))["s"] or 0
            stock_actual = int(entradas - salidas)

            limite = p.stock_min if p.stock_min is not None else threshold
            if stock_actual <= limite:
                productos_bajo_stock.append({
                    "id": p.id,
                    "nombre": p.nombre,
                    "categoria": p.categoria.nombre if p.categoria else None,
                    "stock": stock_actual
                })

        return Response({
            "sucursal": branch.name,
            "fecha_hoy": str(hoy),
            "ventas_hoy": {"monto": ventas_hoy_monto},
            "threshold": threshold,
            "productos_bajo_stock": productos_bajo_stock
        })

    # -------------------------
    # DELETE /api/sucursales/<id>/
    # -------------------------
    def destroy(self, request, *args, **kwargs):
        branch = self.get_object()
        user = request.user

        # Permisos
        if getattr(user, "rol", None) == "limMerchant":
            return Response({"detail": "No puedes borrar sucursales."}, status=403)
        if getattr(user, "rol", None) == "admin" and branch.owner_id != user.id:
            return Response({"detail": "Esta sucursal no te pertenece."}, status=403)

        from django.contrib.auth import get_user_model
        User = get_user_model()

        with transaction.atomic():
            # Borrar empleados limMerchant de esta sucursal
            User.objects.filter(rol="limMerchant", sucursal_id=branch.id).delete()
            # Borrar sucursal
            branch.delete()

        return Response(status=status.HTTP_204_NO_CONTENT)


# =========================
# CATEGORÍAS
# =========================
class CategoryViewSet(viewsets.ModelViewSet):
    serializer_class = CategorySerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ["id", "nombre"]
    search_fields = ["nombre"]
    ordering_fields = ["id", "nombre"]
    ordering = ["id"]

    def get_queryset(self):
        """
        - superuser: ve todo
        - admin: ve categorías donde owner = él
        - limMerchant: ve categorías de su admin (owner de su sucursal)
        """
        user = self.request.user
        qs = Category.objects.all()

        if getattr(user, "is_superuser", False):
            return qs.order_by("id")

        if getattr(user, "rol", None) == "admin":
            return qs.filter(owner=user).order_by("id")

        if getattr(user, "rol", None) == "limMerchant":
            if getattr(user, "sucursal_id", None) and getattr(user, "sucursal", None):
                return qs.filter(owner=user.sucursal.owner).order_by("id")
            return Category.objects.none()

        return Category.objects.none()


# =========================
# PRODUCTOS
# =========================
class ProductViewSet(viewsets.ModelViewSet):
    serializer_class = ProductSerializer
    permission_classes = [permissions.IsAuthenticated]

    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ["id", "categoria", "nombre", "sucursal", "sku"]
    search_fields = ["nombre", "sku"]
    ordering_fields = ["id", "nombre", "precio"]
    ordering = ["id"]

    def get_queryset(self):
        qs = Product.objects.select_related("categoria", "sucursal").all().order_by("id")
        return _scope_by_branch_on_model(qs, self.request.user, branch_field="sucursal")


# =========================
# MOVIMIENTOS DE STOCK
# =========================
class StockMovementViewSet(viewsets.ModelViewSet):
    serializer_class = StockMovementSerializer
    permission_classes = [permissions.IsAuthenticated]

    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ["id", "tipo", "sucursal", "producto"]
    search_fields = ["producto__nombre", "motivo"]
    ordering_fields = ["id", "creado_en"]
    ordering = ["-creado_en"]

    def get_queryset(self):
        qs = (
            StockMovement.objects
            .select_related("producto", "sucursal")
            .all()
            .order_by("-creado_en")
        )
        qs = _scope_by_branch_on_model(qs, self.request.user, branch_field="sucursal")

        # Filtros opcionales por fecha (YYYY-MM-DD)
        desde = self.request.query_params.get("desde")
        hasta = self.request.query_params.get("hasta")
        if desde and hasta:
            try:
                qs = qs.filter(creado_en__date__range=(desde, hasta))
            except Exception:
                pass

        return qs


# =========================
# VENTAS
# =========================
class SaleViewSet(viewsets.ModelViewSet):
    serializer_class = SaleSerializer
    permission_classes = [permissions.IsAuthenticated]

    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ["id", "sucursal"]
    ordering_fields = ["id", "creado_en"]
    ordering = ["-creado_en"]

    def get_queryset(self):
        qs = (
            Sale.objects
            .select_related("sucursal", "usuario")
            .prefetch_related("items__producto")
            .all()
            .order_by("-creado_en")
        )
        return _scope_by_branch_on_model(qs, self.request.user, branch_field="sucursal")


# =========================
# BAJO STOCK (endpoint suelto)
# =========================
@api_view(["GET"])
@permission_classes([permissions.IsAuthenticated])
def low_stock(request):
    """
    Lista productos con stock <= threshold (por movimientos).
    Usa stock_min del producto si está definido; si no, usa 'threshold'.
    Query params:
      - threshold: int (default 5)
      - limit:     int (default 50)
    """
    try:
        threshold = int(request.query_params.get("threshold", 5))
    except ValueError:
        threshold = 5

    try:
        limit = int(request.query_params.get("limit", 50))
    except ValueError:
        limit = 50

    qs = Product.objects.select_related("categoria", "sucursal").all()
    qs = _scope_by_branch_on_model(qs, request.user, branch_field="sucursal")

    rows = []
    for p in qs:
        stock_actual = _stock_actual_producto(p)
        limite = p.stock_min if p.stock_min is not None else threshold
        if stock_actual <= limite:
            rows.append({
                "id": p.id,
                "producto": p.nombre,
                "categoria": p.categoria.nombre if p.categoria else None,
                "sucursal": p.sucursal.name if p.sucursal else None,
                "stock": stock_actual,
            })

    rows = rows[:limit]
    return Response({"threshold": threshold, "items": rows})


# =========================
# EXPORTAR VENTAS A EXCEL (solo admin, con scoping)
# =========================
@api_view(["GET"])
@permission_classes([permissions.IsAuthenticated, IsAdmin])
def export_sales_excel(request):
    # Scope por owner=admin
    qs = (
        Sale.objects
        .prefetch_related("items__producto")
        .select_related("sucursal")
        .all()
    )
    qs = _scope_by_branch_on_model(qs, request.user, branch_field="sucursal")

    wb = Workbook()
    ws = wb.active
    ws.title = "Ventas"
    ws.append(["ID Venta", "Fecha", "Sucursal", "Producto", "Cantidad", "Precio Unit.", "Total Item"])

    for sale in qs:
        for it in sale.items.all():
            ws.append([
                sale.id,
                timezone.localtime(sale.creado_en).strftime("%Y-%m-%d %H:%M"),
                sale.sucursal.name if sale.sucursal else None,
                it.producto.nombre if getattr(it, "producto", None) else None,
                int(it.cantidad),
                float(it.precio_unit),
                float(it.cantidad * it.precio_unit),
            ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="ventas.xlsx"'
    wb.save(response)
    return response


# =========================
# VENTAS HOY (empresa) - JSON
# =========================
@api_view(["GET"])
@permission_classes([permissions.IsAuthenticated])
def ventas_hoy_empresa(request):
    """
    GET /api/ventas/hoy/empresa
    Suma $ de HOY:
      - admin: todas sus sucursales
      - limMerchant: solo su sucursal
      - superuser: todo
    """
    # Día local de hoy
    hoy = timezone.localdate()

    # Ventana del día en TZ local [00:00:00, 23:59:59]
    tz = timezone.get_current_timezone()
    start = timezone.make_aware(datetime.combine(hoy, time.min), tz)
    end   = timezone.make_aware(datetime.combine(hoy, time.max), tz)

    qs = (Sale.objects
          .filter(creado_en__range=(start, end))
          .prefetch_related("items")
          .select_related("sucursal"))

    # Respeta scoping por sucursal/owner
    qs = _scope_by_branch_on_model(qs, request.user, branch_field="sucursal")

    total = 0.0
    for v in qs:
        total += float(sum(it.cantidad * it.precio_unit for it in v.items.all()))

    return Response({"fecha": str(hoy), "monto_total_hoy": total})


# =========================
# KARDEX por producto (JSON + XLSX)
# =========================
@api_view(["GET"])
@permission_classes([permissions.IsAuthenticated])
def kardex_producto(request, producto_id):
    """
    GET /api/productos/<producto_id>/kardex?sucursal=<id>&desde=&hasta=
    Devuelve todos los movimientos de ese producto en esa sucursal,
    ordenados cronológicamente, con saldo acumulado.
    """
    # 1) Validar que venga 'sucursal'
    sucursal_id = request.query_params.get("sucursal")
    if not sucursal_id:
        return Response({"detail": "Parámetro 'sucursal' requerido."}, status=400)

    # 2) Convertir a int (si falla → error)
    try:
        sucursal_id = int(sucursal_id)
    except ValueError:
        return Response({"detail": "Parámetro 'sucursal' inválido."}, status=400)

    # 3) Base QuerySet con producto y sucursal
    qs = (
        StockMovement.objects
        .filter(producto_id=producto_id, sucursal_id=sucursal_id)
        .select_related("producto", "sucursal")
    )

    # 4) Filtros de fecha (opcionales)
    desde = request.query_params.get("desde")  # formato 'YYYY-MM-DD'
    hasta = request.query_params.get("hasta")
    if desde and hasta:
        qs = qs.filter(creado_en__date__range=(desde, hasta))

    # 5) Scoping por rol / empresa
    qs = _scope_by_branch_on_model(qs, request.user, branch_field="sucursal")

    # 6) Orden cronológico
    qs = qs.order_by("creado_en")

    # 7) Construir respuesta con saldo acumulado
    saldo = 0
    items = []
    for m in qs:
        delta = m.cantidad if m.tipo == "IN" else -m.cantidad
        saldo += delta
        items.append({
            "id": m.id,
            "fecha": m.creado_en,   # si querés, después te lo paso en zona horaria local
            "tipo": m.tipo,         # IN / OUT
            "cantidad": m.cantidad,
            "motivo": m.motivo,
            "saldo": saldo,         # saldo luego de este movimiento
        })

    # 8) Respuesta final
    return Response({
        "producto_id": int(producto_id),
        "sucursal_id": sucursal_id,
        "items": items
    })


@api_view(["GET"])
@permission_classes([permissions.IsAuthenticated])
def kardex_producto_xlsx(request, producto_id):
    sucursal_id = request.query_params.get("sucursal")
    if not sucursal_id:
        return Response({"detail": "Parámetro 'sucursal' requerido."}, status=400)
    qs = (StockMovement.objects.filter(producto_id=producto_id, sucursal_id=sucursal_id)
          .select_related("producto", "sucursal").order_by("creado_en"))
    d = request.query_params.get("desde"); h = request.query_params.get("hasta")
    if d and h: qs = qs.filter(creado_en__date__range=(d, h))
    qs = _scope_by_branch_on_model(qs, request.user, branch_field="sucursal")

    wb = Workbook(); ws = wb.active
    ws.title = "Kardex"
    ws.append(["Fecha","Tipo","Cantidad","Motivo","Saldo"])
    saldo = 0
    for m in qs:
        delta = m.cantidad if m.tipo == "IN" else -m.cantidad
        saldo += delta
        ws.append([
            timezone.localtime(m.creado_en).strftime("%Y-%m-%d %H:%M"),
            m.tipo, int(m.cantidad), m.motivo or "", int(saldo)
        ])
    resp = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp['Content-Disposition'] = f'attachment; filename="kardex_producto_{producto_id}_suc_{sucursal_id}.xlsx"'
    wb.save(resp); return resp
